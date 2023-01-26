#!/usr/bin/python3
import os
import sys
import shutil
from PIL import Image

#using extendedopenpyxl as it saves correctly when a single cell contains 2 fonts in an excel file, normal openpyxl DOES NOT
from extendedopenpyxl import load_workbook, save_workbook
from openpyxl.drawing.image import Image as Im
from pdf2image import convert_from_path
import time

# about this module pdf2image that converts pdf to images needs files and this function only works if you explicitly state where the bin files is at.....
# therefore must save the files to a server (or some known location) and call the path to the bin file in the poppler path called out later, that is
# if other people want to use the program

import easygui
from PyQt5.QtWidgets import QApplication, QLabel, QDialog, QLineEdit, QCheckBox, QPushButton
from PyQt5.QtCore import Qt, pyqtSignal, QThread
from PyQt5.QtGui import QFont
import getpass

#========================
#HOW PROGRAM WORKS:
#========================
   #The program will display a GUI and when you hit Start, it will attempt find the file name PDF that is entered into the GUI
   #at a specific location that is in the "ballooned_folder" variable.  If it finds the filename, it will convert the PDF to
   #an image, then does smart cropping by reading background RGB value and iterating through the top/sides/bottom to find RGB value
   #that is different.  Once it finds RGB values different, sets those areas for cropping.  The program then resizes after cropping
   #and inserts the images into an excel file you specify.  It's setup to enter the image(s) into the excel in a specific way for myself.


def process(file):
    #tolerance so that RGB value can have some variability before it's too far out to find the cropping spot
    TOLERANCE = 11

    #get size of image file
    image = Image.open(file)
    image_width = image.size[0]
    image_height = image.size[1]
    rgb_image = image.convert('RGB')

    #Sample background color
    def rgb_tuple_to_str(tuple):
        return 'rgb(' + str(tuple[0]) + ', ' + str(tuple[1]) + ', ' + str(tuple[2]) + ')'

    #break color RGB down to check against background RGB if it's outside the background RGB values or not (with tolerance)
    def is_like_bg_color(color):
        color_r, color_g, color_b = color[0], color[1], color[2]
        bg_r, bg_g, bg_b = bg_color[0], bg_color[1], bg_color[2]
        r_similar, g_similar, b_similar = False, False, False  

        if color_r in range(bg_r - TOLERANCE, bg_r + TOLERANCE):
            r_similar = True
            
        if color_g in range(bg_g - TOLERANCE, bg_g + TOLERANCE):
            g_similar = True
            
        if color_b in range(bg_b - TOLERANCE, bg_b + TOLERANCE):
            b_similar = True

        return r_similar and g_similar and b_similar


    #how far in the corners where to sample background color
    x_offset = image_width * 0.001
    y_offset = image_height * 0.001

    #sample background color in all 4 corners of image to make sure background color matches, must match to
    #make sure you have the correct background color
    ul_color = rgb_image.getpixel((x_offset, y_offset))
    ur_color = rgb_image.getpixel((image_width - x_offset, y_offset))
    ll_color = rgb_image.getpixel((x_offset, image_height - y_offset))
    lr_color = rgb_image.getpixel((image_width - x_offset, image_height - y_offset))
    bg_color = ()

    #check to make sure background color matches to make sure you have the true background color
    if ul_color == ur_color and ur_color == ll_color and ll_color == lr_color:
        bg_color = ul_color

    #Iterates top edge pixels looking where the RGB value changes from background color and sets top edge for crop spot
    #set to back off by -3 pixels from where the RGB value changes, set that to whatever you want
    #NOTE: i have this set to ignore the top left edge for my needs by starting it at 850 pixel instead of 0 pixel

    top_edge_coords = []

    for i in range(850, image_width, int(image_height / 10)):
        for y in range(1, image_height - 1):
            if not is_like_bg_color(rgb_image.getpixel((i, y))):
                top_edge_coords.append(y - 3)
                break

    top_edge_coord = top_edge_coords[0]
    for c in top_edge_coords:
        if c < top_edge_coord:
            top_edge_coord = c

    #Iterates bottom edge pixels of where the RGB value changes from background color and sets bottom edge for crop spot
    #set to back off by +4 pixels from where the RGB value changes, set that to whatever you want

    bottom_edge_coords = []

    for i in range(0, image_width, int(image_height / 10)):
        for y in range(image_height - 1, 0, -1):
            if not is_like_bg_color(rgb_image.getpixel((i, y))):
                bottom_edge_coords.append(y + 4)
                break

    bottom_edge_coord = bottom_edge_coords[0]
    for c in bottom_edge_coords:
        if c > bottom_edge_coord:
            bottom_edge_coord = c

    #Iterates left edge pixels of where the RGB value changes from background color and sets left edge for crop spot
    #set to back off by -3 pixels from where the RGB value changes, set that to whatever you want
    #NOTE: i have this set to ignore the top left edge for my needs by starting it at 120th pixel instead of 0 pixel

    left_edge_coords = []

    for i in range(120, image_height, 4):
        for x in range(0, image_width - 1):
            if not is_like_bg_color(rgb_image.getpixel((x, i))):
                left_edge_coords.append(x - 3)
                break

    left_edge_coord = left_edge_coords[0]
    for c in left_edge_coords:
        if c < left_edge_coord:
            left_edge_coord = c

    #Iterates right edge pixels of where the RGB value changes from background color and sets right edge for crop spot
    #set to back off by +4 pixels from where the RGB value changes, set that to whatever you want

    right_edge_coords = []

    for i in range(0, image_height, 4):
        for x in range(image_width - 1, 0, -1):
            try:
                if not is_like_bg_color(rgb_image.getpixel((x, i))):
                    right_edge_coords.append(x + 4)
                    break
            except IndexError:
                pass

    right_edge_coord = right_edge_coords[0]
    for c in right_edge_coords:
        if c > right_edge_coord:
            right_edge_coord = c

    #Crop image
    cropped_image = image.crop((left_edge_coord, top_edge_coord, right_edge_coord, bottom_edge_coord))

    #save image
    cropped_image.save(file)


class Actions(QDialog):

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):

        #create UI
        self.setGeometry(400,200,350,120)
        self.setFixedSize(self.size())
        self.setWindowFlag(Qt.WindowMinimizeButtonHint, True)

        self.setWindowTitle("Program Name")

        self.about = QLabel("About: Section", self)
        self.about.adjustSize()
        self.about.move(25, 0)

        self.instruct = QLabel("---  Some Instructions for use", self)
        self.instruct.setFont(QFont('Arial', 7))
        self.instruct.adjustSize()
        self.instruct.move(55, 20)

        self.part_label = QLabel("<b>Button Label:<B>", self)
        self.part_label.adjustSize()
        self.part_label.move(15, 35)

        self.part = QLineEdit(self)
        self.part.setFixedWidth(100)
        self.part.move(15, 50)

        self.check = QCheckBox(self)
        self.check.setText("Open Excel When Done?")
        self.check.move(135, 50)

        self.button = QPushButton(self)
        self.button.setText("Start")
        self.button.clicked.connect(self.onButtonClick)
        self.button.move(15, 80)

        self.update_label = QLabel("update labeling                                                                                                  ", self)
        self.update_label.adjustSize()
        self.update_label.move(110, 85)

        self.show()

    def onButtonClick(self):
        try:
            #remove any unnecessary spaces from name of file entered in and make it all upper case
            a = self.part.text().strip(" ")
            a_upper = a.upper()

            #check state of checkbox
            b = self.check.checkState()

            #default finding excel file to copy the images to , to false, if finds excel, sets to true later
            path_exists = False

            #update GUI with what program is doing
            self.update_label.setText("Starting.....finding files....")

            #custom error check to make sure file name entered into field is exactly 12 characters
            #change this or remove it if you don't need it
            if len(a) != 12:
                msg = easygui.msgbox("Custom Error message", "ERROR")
                self.update_label.setText("Try Again....")
            else:
                #call find_excel function to look for excel file and get it's file path
                excel_path = self.find_excel(a_upper)

                #if file path of excel doesn't exist, display error message
                if not os.path.exists(excel_path):
                    msg = easygui.msgbox("Custom Error message if file not found", "ERROR")
                    self.update_label.setText("Try Again....")
                else:
                    #if finds excel, set path exists to true
                    path_exists = True

            #above code tests if FAI excel and directory exists

            if path_exists == True:

                #enter base file path location for PDF location for variable
                ballooned_folder = "O:\\......"

                #read file path
                read_ballooned_folder = os.listdir(ballooned_folder)

                ballooned_rev_check = False
                ballooned_pdf_check = False

                #iterate file path to find the PDF filename entered into the line box on the GUI
                for i in read_ballooned_folder:
                    if a_upper in i:
                        #set value to true if PDF filename exists in folder
                        ballooned_rev_check = True
                        if ".pdf" == i[-4:]:
                            #set to true if file is actually a PDF extension
                            ballooned_pdf_check = True
                            #get complete file path location
                            ballooned_path = ballooned_folder + "\\" + i

                #code checks if file, rev and pdf exists in ballooned folder

                if ballooned_rev_check == True:
                    if ballooned_pdf_check == True:
                        #if file name and pdf extension found, start cropping and doing work starting with class External
                        self.calc = External(a_upper, b, ballooned_path, excel_path)
                        self.calc.updateChanged.connect(self.onupdateChanged)
                        self.calc.exitChanged.connect(self.onexitChanged)
                        self.calc.start()
                    else:
                        msg = easygui.msgbox("ERROR if PDF extension", "ERROR")
                        self.update_label.setText("Try Again....")
                else:
                    msg = easygui.msgbox("Error if filename not found in folder", "ERROR")
                    self.update_label.setText("Try Again....")

            #above code if file exists execute program if not output errors based on results if no rev or pdf
        except:
            easygui.msgbox(msg="ERROR: 6006", title="ERROR")

    def find_excel(self, a):
        #use this method to come up with the file path to find your excel
        #this is a custom method needed for my needs to get the correct excel path
        #change this to whatever you need, i had more code here i deleted that was just very specific to my needs

        #add excel extension to filename entered into GUI
        partexcel = a + ".xlsx"
        #enter your file path where the excel should exist here
        folderfai = "O:\\......"

        return(folderfai)


    def onupdateChanged(self, value):
        #for updated label in GUI to update what program is doing
        self.update_label.setText(value)

    def onexitChanged(self, value1):
        #if program done, terminate Qthread
        if "Yes" in value1:
            self.calc.terminate()

    #terminate of thread not used, but keeping code in here in case i want to use it later

class External(QThread):
    updateChanged = pyqtSignal(str)
    exitChanged = pyqtSignal(str)

    def __init__(self, a, b, ballooned_path, excel_path):
        super(External, self).__init__()
        self.a = a
        self.b = b
        self.ballooned_path = ballooned_path
        self.excel_path = excel_path

    def run(self):
        import win32com.client
        import pythoncom
        pythoncom.CoInitialize()

        #import statements needs to be called in this thread rather than top of program or the api calls don't work
        #i don't know why

        try:
            self.updateChanged.emit("Converting Ballooned PDF to Images.....")

            #NOTE: using user name to find correct path for INPUT_DIR, TEMP_EXCEL, you may not need this!
            user = getpass.getuser()

            #enter temporary input location for cropped files
            INPUT_DIR = "O:\\..."
            #enter temporary location to save a temporary excel
            TEMP_EXCEL = "O:\\..."

            # Create temporary input directory, if not present, folder will be deleted later
            try:
                os.stat(INPUT_DIR)
            except:
                os.mkdir(INPUT_DIR)

            try:
                os.stat(TEMP_EXCEL)
            except:
                os.mkdir(TEMP_EXCEL)

            #removes any files in directory in case any exists if the dir exists, needs a clean directory
            for file in os.listdir(INPUT_DIR):
                os.remove(os.path.join(INPUT_DIR, file))

            #ENTER POPPLER PATH here, THIS IS A MUST SO THAT PDF CAN BE CONVERTED TO IMAGE
            images = convert_from_path(self.ballooned_path, poppler_path="O:........\\poppler-22.04.0\\Library\\bin")

            #module that converts pdf to images needs files and this function only works if you explicitly state where the bin files is at.....
            #therefore must save the files to specific location and call the path to the bin file here if other people want to use the program

            #update GUI with which page of PDF currently being cropped
            for i in range(len(images)):
                self.updateChanged.emit("Converting Ballooned PDF page " + str(i) + " to Image.....")
                images[i].save(INPUT_DIR + "\\" + "page" + str(i) + ".png", "PNG")


            #Iterate over working directory in order
            list_dir_to_read = sorted(os.listdir(INPUT_DIR), key=len)

            for file in sorted(os.listdir(INPUT_DIR), key=len):
                file_index = list_dir_to_read.index(file) + 1
                self.updateChanged.emit("Cropping Ballooned page " + str(file_index) + " .....")

                file_path = os.path.join(INPUT_DIR, file)

                #call process function and process image for cropping
                process(file_path)

            #iterates through images in temporary image folder and runs process function to crop them

            wb = load_workbook(self.excel_path, keep_links=False, keep_vba= True)
            all_sheets = wb.sheetnames

            #deletes excel sheets that have a "D"
            #NOTE: THIS is VERY specific to my needs, remove this or change it

            drawing_pages = [x for x in all_sheets if "D" in x.upper()]
            for i in drawing_pages:
                del wb[i]
            #deletes all drawing pages from excel to be re-created later

            read_images = sorted(os.listdir(INPUT_DIR), key=len)
            clean_images_folder_list = [x for x in read_images if ".png" in x]
            #needs above line of code to remove any windows thumb.db files that get created inthe folder

            total_images = len(clean_images_folder_list)

            #iterate all images to rotate/adding new sheets to excel file and adding images
            k = 0
            while k < total_images:
                page = k + 1

                #update GUI with which image getting added to excel
                self.updateChanged.emit("Resizing & Adding Drawing Page " + str(page) + " to Excel.....")
                image_path = INPUT_DIR + "\\" + clean_images_folder_list[k]


                img_size = Image.open(image_path)
                rotated = img_size.transpose(Image.Transpose.ROTATE_270)
                rotated.save(image_path)
                # above, using PIL Image to rotate and resaving

                logo = Im(image_path)

                #change image to specific width
                logo.height = 1000
                logo.width = 700
                # above using Image from openpyxl to resize... if trying to resize using PIL Image... image quality degrades, but doing
                # it with openpyxl Image doesn't.... i don't know why.

                #add new sheet to excel
                wb.create_sheet("D" + str(k + 1))
                ws = wb.get_sheet_by_name("D" + str(k + 1))

                #set sheet margins
                ws.page_margins.left = 0
                ws.page_margins.right = 0
                ws.page_margins.top = 0
                ws.page_margins.bottom = 0
                ws.page_margins.footer = 0
                ws.page_margins.header = 0

                #where to add image
                ws.add_image(logo, "A1")
                k += 1

            #update GUI
            self.updateChanged.emit("Saving Excel!")


            #Save Excel at a specific location temporarily as a XLSM (THIS IS VERY SPECIFIC NEEDS)
            #i temporarily save it as a XLSM and then use Excel API later to resave
            #it as a xlsx, this must be done for me so things like checkboxes don't get remove
            #from the excel file.  You may not need to do this for your needs.

            part_number = self.excel_path[-17:]
            part_number_change_ext = part_number.replace(".xlsx", ".xlsm")
            TEMP_EXCEL_FULL_PATH = TEMP_EXCEL + "\\" + part_number_change_ext
            save_workbook(wb, TEMP_EXCEL_FULL_PATH)

            time.sleep(2)

            #save to xlsm to temp folder to keep the checkboxes on the excel... checkboxes will be deleted
            #if saved as an xlsx, using excel API to convert the xlsm to xlsx and save in appropriate spot
            #afterwards.  saving using excel api doesn't remove the checkboxes

            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            data = excel.Workbooks.Open(TEMP_EXCEL_FULL_PATH)

            #open xlsm to save as xlsx using excel api
            data.CheckCompatibility = False
            data.DoNotPromptForConvert = True

            #removes alerts and compatability messages
            try:
                data.SaveAs(self.excel_path, FileFormat=51)
                data.Close()

                #save and close using the excel API, fileformat 51 is xlsx
                shutil.rmtree(INPUT_DIR)
                shutil.rmtree(TEMP_EXCEL)
                self.updateChanged.emit("Complete!")
                if int(self.b) == 2:
                    os.system(f'start EXCEL.EXE "{self.excel_path}"')
            except:
                data.Close()
                shutil.rmtree(INPUT_DIR)
                shutil.rmtree(TEMP_EXCEL)
                self.updateChanged.emit("ERROR SAVING!")
                easygui.msgbox("Save & Close excel to FAI folder, then start!", "ERROR")

        except:
            easygui.msgbox(msg = "ERROR: nub", title="ERROR")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = Actions()
    sys.exit(app.exec_())
